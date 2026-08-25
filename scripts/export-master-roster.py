#!/usr/bin/env python3
"""Build the master roster workbook.

Two sources are merged:

  * data/roster-history.json - the parsed original workbook (Mar-Aug 2026).
    This is the archive: it is never edited again.
  * the live artifact state    - whatever the roster page holds right now.
    It wins wherever the two overlap, because that is what the admin last
    published.

Usage:
    python3 scripts/export-master-roster.py \
        --state artifact-state.json \
        [--history data/roster-history.json] \
        [--out "Shan Village - Master Roster.xlsx"]
"""
import argparse, datetime, json, os, re, sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - environment problem, not a code path
    sys.exit("openpyxl is required: pip install openpyxl")

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]

# ----------------------------------------------------------------- time --
def mins(t):
    h, m = t.split(":")
    return int(h) * 60 + int(m)

def seg(a, b, cross):
    m = mins(b) - mins(a)
    if cross or m <= 0:
        m += 1440
    return m

def hours_of(c):
    """Mirrors hoursOf() in the roster page. Keep the two in step."""
    if c.get("k") != "WORK" or not c.get("s"):
        return 0.0
    if c.get("sp") and c.get("s2") and c.get("e2"):
        m = seg(c["s"], c["e"], False) + seg(c["s2"], c["e2"], bool(c.get("x")))
    else:
        m = seg(c["s"], c["e"], bool(c.get("x")))
    return round(m / 60, 2)

def end_label(t, cross):
    return "24:00" if cross and t == "00:00" else t

def cell_label(c):
    k = c.get("k")
    if k == "OTHER":
        return c.get("note") or "Other duty"
    if k != "WORK":
        return {"OFF": "OFF", "PH": "Public holiday", "LEAVE": "Leave"}.get(k, k or "")
    if not c.get("s"):
        return "On duty"
    if c.get("sp") and c.get("s2") and c.get("e2"):
        return "%s-%s / %s-%s" % (c["s"], c["e"], c["s2"], end_label(c["e2"], c.get("x")))
    return "%s-%s" % (c["s"], end_label(c["e"], c.get("x")))

def parse_label(label):
    """'09:00-14:00 / 19:00-24:00' -> cell fields. Accepts - or en dash."""
    parts = [p.strip() for p in label.split("/")]
    segs = []
    for part in parts:
        a, b = re.split(r"[-–—]", part, maxsplit=1)
        segs.append((a.strip(), b.strip()))

    def norm(t):
        h, m = t.split(":")
        h, m = int(h), int(m)
        wrapped = h >= 24
        if wrapped:
            h -= 24
        return "%02d:%02d" % (h, m), wrapped

    if len(segs) == 1:
        s, e = segs[0]
        e_, wrapped = norm(e)
        return {"k": "WORK", "s": s, "e": e_,
                "x": wrapped or mins(e_) <= mins(s),
                "sp": False, "s2": "", "e2": ""}
    s, e = segs[0]
    s2, e2 = segs[1]
    e_, _ = norm(e)
    e2_, wrapped = norm(e2)
    return {"k": "WORK", "s": s, "e": e_,
            "x": wrapped or mins(e2_) <= mins(s2),
            "sp": True, "s2": s2, "e2": e2_}

def day_date(week_start, i):
    return datetime.date.fromisoformat(week_start) + datetime.timedelta(days=i)

# ----------------------------------------------------------------- load --
def load_history(path):
    """-> (weeks{start: {name: {day: cell}}}, staff_meta{name: (pos, outlet)})"""
    if not path or not os.path.exists(path):
        return {}, {}
    raw = json.load(open(path, encoding="utf8"))
    people, weeks = raw["people"], raw["weeks"]
    out, meta = {}, {}
    for name, pos, outlet in people:
        meta[name] = (pos or "", outlet or "")
    for row in raw["rows"]:
        wi, pi, di, kind, label = row[0], row[1], row[2], row[3], row[4]
        name = people[pi][0]
        start = weeks[wi][0]
        if kind == "WORK" and re.search(r"[-–—]", label) and ":" in label:
            cell = parse_label(label)
        elif kind in ("WORK", "TRIAL"):
            cell = {"k": "WORK"}
        elif kind in ("OFF", "PH", "LEAVE"):
            cell = {"k": kind}
        else:
            # anything the original sheet spelled out in words, e.g. "Visa Extend"
            cell = {"k": "OTHER", "note": row[6] or None}
        raw = row[8] if len(row) > 8 else ""
        if isinstance(raw, str) and raw.strip():
            cell["raw"] = raw.strip()
        out.setdefault(start, {}).setdefault(name, {})[di] = cell
    return out, meta

def load_state(path):
    raw = json.load(open(path, encoding="utf8"))
    if "roster" not in raw:          # a whole page was handed over, not the state
        m = re.search(r'<script id="state" type="application/json">(\{.*?)</script>',
                      open(path, encoding="utf8").read(), re.S)
        if not m:
            sys.exit("no roster state found in %s" % path)
        raw = json.loads(m.group(1))
    return raw

# --------------------------------------------------------------- styles --
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="171310")
HEAD_FONT = Font(color="F7EFDD", bold=True, size=11)
SUB_FILL = PatternFill("solid", fgColor="F3ECDC")
OFF_FILL = PatternFill("solid", fgColor="ECE5D5")
PH_FILL = PatternFill("solid", fgColor="E1F0E4")
LEAVE_FILL = PatternFill("solid", fgColor="F6ECD3")
OT_FILL = PatternFill("solid", fgColor="F7E1DA")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

def header_row(ws, row, values, widths=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill, c.font, c.alignment, c.border = HEAD_FILL, HEAD_FONT, CENTER, BORDER
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def fill_for(cell):
    return {"OFF": OFF_FILL, "PH": PH_FILL, "LEAVE": LEAVE_FILL}.get(cell.get("k"))

# ----------------------------------------------------------------- main --
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--state", required=True)
    ap.add_argument("--history", default="data/roster-history.json")
    ap.add_argument("--out", default="Shan Village - Master Roster.xlsx")
    ap.add_argument("--recent", type=int, default=0,
                    help="weekly grid covers only the most recent N weeks "
                         "(monthly totals still cover every month)")
    ap.add_argument("--compact", action="store_true",
                    help="leave out the per-day Shift detail sheet (much smaller file)")
    ap.add_argument("--generated-at", default=None,
                    help="ISO timestamp for the cover sheet (default: now, UTC)")
    args = ap.parse_args()

    state = load_state(args.state)
    hist, hist_meta = load_history(args.history)

    limit = state.get("ot")
    limit = float(limit) if isinstance(limit, (int, float)) and limit > 0 else 10.0

    # staff: the artifact is the roll of who is here now; history adds leavers
    staff = []          # (name, position, outlet, current)
    seen = set()
    for s in state.get("staff", []):
        name = s.get("name", "").strip()
        if not name or name in seen:
            continue
        seen.add(name)
        staff.append((name, s.get("pos", ""), hist_meta.get(name, ("", ""))[1], True))
    for name, (pos, outlet) in sorted(hist_meta.items()):
        if name not in seen:
            seen.add(name)
            staff.append((name, pos, outlet, False))
    pos_of = {n: p for n, p, o, c in staff}

    # merge: history first, live state on top
    weeks = {}
    for start, rows in hist.items():
        for name, days in rows.items():
            for di, cell in days.items():
                weeks.setdefault(start, {}).setdefault(name, {})[int(di)] = cell
    id_name = {s["id"]: s.get("name", "") for s in state.get("staff", [])}
    for start, rows in (state.get("roster") or {}).items():
        for pid, days in rows.items():
            name = id_name.get(pid)
            if not name:
                continue
            for di, cell in days.items():
                weeks.setdefault(start, {}).setdefault(name, {})[int(di)] = cell

    week_starts = sorted(weeks)
    gen = args.generated_at or datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")

    wb = Workbook()

    # ---------------------------------------------------------- cover ---
    ws = wb.active
    ws.title = "Read me"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 78
    ws["A1"] = "SHAN VILLAGE - MASTER DUTY ROSTER"
    ws["A1"].font = Font(bold=True, size=16)
    facts = [
        ("Generated", gen + "  (automatic daily export)"),
        ("Roster last published", state.get("pub") or "not published yet"),
        ("Weeks covered", "%s to %s (%d weeks)" % (
            week_starts[0], str(day_date(week_starts[-1], 6)), len(week_starts)) if week_starts else "none"),
        ("People on file", "%d (%d currently on the roster)" % (
            len(staff), sum(1 for s in staff if s[3]))),
        ("Overtime rule", "hours above %g in a single day count as overtime" % limit),
        ("", ""),
        ("Sheet: Staff", "everyone on file, position, and whether they are still on the roster"),
        ("Sheet: Monthly overtime", "per person per month - normal, overtime, total, days over the limit"),
        ("Sheet: Weekly grid", "the familiar week-by-week layout, newest week first"),
        ("Sheet: Shift detail", "one row per person per day - use this for pivot tables"),
        ("Sheet: Change log", "who changed what in the roster page, and when"),
        ("", ""),
        ("Source", "history from the original duty-roster workbook; everything from"),
        ("", "the live roster page from the point it took over. The page wins on overlap."),
    ]
    r = 3
    for k, v in facts:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=False)
        r += 1

    # ---------------------------------------------------------- staff ---
    ws = wb.create_sheet("Staff")
    header_row(ws, 1, ["Employee", "Position", "Outlet", "On the roster now"],
               [26, 22, 16, 18])
    for i, (name, pos, outlet, current) in enumerate(staff, start=2):
        ws.cell(row=i, column=1, value=name).border = BORDER
        ws.cell(row=i, column=2, value=pos or "-").border = BORDER
        ws.cell(row=i, column=3, value=outlet or "-").border = BORDER
        c = ws.cell(row=i, column=4, value="Yes" if current else "No (past)")
        c.border, c.alignment = BORDER, CENTER
    ws.freeze_panes = "A2"

    # -------------------------------------------------- shift detail ---
    detail = []          # (date, name, kind, label, hours, normal, ot)
    for start in week_starts:
        for name, days in weeks[start].items():
            for di in range(7):
                cell = days.get(di)
                if not cell:
                    continue
                h = hours_of(cell)
                detail.append((day_date(start, di), name, cell, h,
                               min(h, limit), max(0.0, round(h - limit, 2))))
    detail.sort(key=lambda x: (x[0], x[1]))

    ws = wb.create_sheet("Shift detail")
    if args.compact:
        wb.remove(ws)
        ws = wb.create_sheet("_tmp")
    header_row(ws, 1, ["Date", "Day", "Week starting", "Employee", "Position",
                       "Status", "Shift", "Start", "End", "2nd start", "2nd end",
                       "Hours", "Normal", "Overtime", "As written originally"],
               [12, 6, 14, 24, 20, 12, 22, 8, 8, 9, 9, 8, 8, 9, 20])
    for i, (d, name, cell, h, normal, ot) in enumerate(detail, start=2):
        vals = [d.isoformat(), DAYS[d.weekday()],
                (d - datetime.timedelta(days=d.weekday())).isoformat(),
                name, pos_of.get(name) or "", cell.get("k", ""), cell_label(cell),
                cell.get("s", ""), end_label(cell.get("e", ""), cell.get("x")) if cell.get("e") else "",
                cell.get("s2", ""), end_label(cell.get("e2", ""), cell.get("x")) if cell.get("e2") else "",
                h or "", normal or "", ot or "", cell.get("raw", "")]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v)
        if ot:
            ws.cell(row=i, column=14).fill = OT_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:O%d" % max(1, len(detail) + 1)

    if args.compact:
        wb.remove(wb["_tmp"])

    # ----------------------------------------------- monthly overtime ---
    months = {}
    for d, name, cell, h, normal, ot in detail:
        key = (d.year, d.month)
        m = months.setdefault(key, {}).setdefault(
            name, {"normal": 0.0, "ot": 0.0, "worked": 0, "off": 0, "leave": 0,
                   "ph": 0, "over_days": 0})
        m["normal"] += normal
        m["ot"] += ot
        k = cell.get("k")
        if k == "WORK":
            m["worked"] += 1
        elif k == "OFF":
            m["off"] += 1
        elif k == "LEAVE":
            m["leave"] += 1
        elif k == "PH":
            m["ph"] += 1
        if ot:
            m["over_days"] += 1

    ws = wb.create_sheet("Monthly overtime")
    ws["A1"] = "Overtime counted above %g hours in a day" % limit
    ws["A1"].font = Font(bold=True, size=12)
    row = 3
    for (y, mth) in sorted(months):
        ws.cell(row=row, column=1, value="%s %d" % (MONTHS[mth - 1], y)).font = Font(bold=True, size=12)
        row += 1
        header_row(ws, row, ["Employee", "Position", "Normal hours", "Overtime hours",
                             "Total hours", "Days worked", "Days off", "Leave",
                             "Public hol.", "Days over %g h" % limit],
                   [26, 20, 13, 14, 12, 12, 10, 8, 11, 14])
        row += 1
        tot = {"normal": 0.0, "ot": 0.0, "worked": 0, "off": 0, "leave": 0, "ph": 0, "over_days": 0}
        for name in sorted(months[(y, mth)], key=lambda n: [s[0] for s in staff].index(n)
                           if n in [s[0] for s in staff] else 999):
            m = months[(y, mth)][name]
            for k in tot:
                tot[k] += m[k]
            vals = [name, pos_of.get(name) or "-", round(m["normal"], 2), round(m["ot"], 2),
                    round(m["normal"] + m["ot"], 2), m["worked"], m["off"], m["leave"],
                    m["ph"], m["over_days"]]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=j, value=v)
                c.border = BORDER
                if j >= 3:
                    c.alignment = CENTER
            if m["ot"]:
                ws.cell(row=row, column=4).fill = OT_FILL
            row += 1
        vals = ["Team total", "", round(tot["normal"], 2), round(tot["ot"], 2),
                round(tot["normal"] + tot["ot"], 2), tot["worked"], tot["off"],
                tot["leave"], tot["ph"], tot["over_days"]]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.border, c.font, c.fill = BORDER, Font(bold=True), SUB_FILL
            if j >= 3:
                c.alignment = CENTER
        row += 3

    # ----------------------------------------------------- weekly grid ---
    ws = wb.create_sheet("Weekly grid")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    for i in range(3, 10):
        ws.column_dimensions[get_column_letter(i)].width = 19
    ws.column_dimensions["J"].width = 10
    row = 1
    grid_weeks = list(reversed(week_starts))
    if args.recent and args.recent > 0:
        dropped = grid_weeks[args.recent:]
        grid_weeks = grid_weeks[:args.recent]
        if dropped:
            # Say what is not here. A grid that quietly stops at eight weeks
            # reads as "that is all there ever was".
            c = ws.cell(row=row, column=1,
                        value="Weekly grid shows the most recent %d weeks. "
                              "%d earlier weeks (from %s) are in the full export and "
                              "are still counted in Monthly overtime."
                              % (len(grid_weeks), len(dropped), min(dropped)))
            c.font = Font(italic=True, color="756B58")
            row += 2
    for start in grid_weeks:
        dates = [day_date(start, i) for i in range(7)]
        c = ws.cell(row=row, column=1,
                    value="Week %s to %s" % (dates[0].strftime("%d %b %Y"),
                                             dates[6].strftime("%d %b %Y")))
        c.font = Font(bold=True, size=13)
        row += 1
        header_row(ws, row, ["Employee", "Position"] +
                   ["%s %s" % (DAYS[i], dates[i].strftime("%d %b")) for i in range(7)] +
                   ["Hours"])
        row += 1
        present = [s for s in staff if s[0] in weeks[start]]
        for name, pos, outlet, current in present:
            days = weeks[start][name]
            ws.cell(row=row, column=1, value=name).border = BORDER
            ws.cell(row=row, column=2, value=pos or "-").border = BORDER
            total = 0.0
            for i in range(7):
                cell = days.get(i)
                c = ws.cell(row=row, column=3 + i, value=cell_label(cell) if cell else "")
                c.border, c.alignment = BORDER, CENTER
                if cell:
                    f = fill_for(cell)
                    if f:
                        c.fill = f
                    h = hours_of(cell)
                    total += h
                    if h > limit:
                        c.font = Font(bold=True, color="9C2B1F")
            c = ws.cell(row=row, column=10, value=round(total, 2))
            c.border, c.alignment, c.font = BORDER, CENTER, Font(bold=True)
            row += 1
        row += 2
    ws.freeze_panes = "C1"

    # ------------------------------------------------------ change log ---
    ws = wb.create_sheet("Change log")
    header_row(ws, 1, ["When", "Who", "What changed"], [22, 12, 110])
    log = state.get("log") or []
    for i, entry in enumerate(log, start=2):
        ws.cell(row=i, column=1, value=entry.get("at", "")).border = BORDER
        ws.cell(row=i, column=2, value=(entry.get("who", "") or "").title()).border = BORDER
        ws.cell(row=i, column=3, value=entry.get("what", "")).border = BORDER
    if not log:
        ws.cell(row=2, column=1, value="No changes recorded yet.")
    ws.freeze_panes = "A2"

    wb.save(args.out)
    print("wrote %s  (%d weeks, %d people, %d shift rows, %d months)"
          % (args.out, len(week_starts), len(staff), len(detail), len(months)))

if __name__ == "__main__":
    main()
