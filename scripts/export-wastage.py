#!/usr/bin/env python3
"""Build the daily wastage workbook from the wastage page's state.

    python3 scripts/export-wastage.py --state <artifact.html|state.json> \
        [--day 2026-08-25] [--out "Shan Village Wastage 2026-08-25.xlsx"]

--day defaults to today in Abu Dhabi. Pass --all to write every day the
page holds rather than one day.
"""
import argparse, csv, datetime, json, re, sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

GULF = datetime.timezone(datetime.timedelta(hours=4))

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="171310")
HEAD_FONT = Font(color="F7EFDD", bold=True, size=11)
SUB_FILL = PatternFill("solid", fgColor="F3ECDC")
MONEY_FILL = PatternFill("solid", fgColor="F7E1DA")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(vertical="top", wrap_text=True)


def load_state(path):
    raw = open(path, encoding="utf8").read()
    m = re.search(r'<script id="state" type="application/json">(\{.*?)</script>', raw, re.S)
    text = m.group(1) if m else raw
    state = json.loads(text)
    if "entries" not in state:
        sys.exit("no wastage entries found in %s" % path)
    return state


def header(ws, row, values, widths=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill, c.font, c.alignment, c.border = HEAD_FILL, HEAD_FONT, CENTER, BORDER
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def num(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--state", required=True)
    ap.add_argument("--day")
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--out")
    ap.add_argument("--generated-at")
    ap.add_argument("--csv", action="store_true",
                    help="print the rows as CSV on stdout instead of writing a workbook. "
                         "Text this size crosses into Google Drive intact; a binary "
                         "workbook does not.")
    args = ap.parse_args()

    state = load_state(args.state)
    currency = state.get("cur") or "AED"
    entries = state.get("entries") or []

    day = args.day or datetime.datetime.now(GULF).strftime("%Y-%m-%d")
    rows = entries if args.all else [e for e in entries if e.get("d") == day]
    rows.sort(key=lambda e: (e.get("d") or "", e.get("t") or ""))

    out = args.out or ("Shan Village Wastage %s.xlsx" % ("all days" if args.all else day))
    gen = args.generated_at or datetime.datetime.now(GULF).strftime("%Y-%m-%d %H:%M")

    if args.csv:
        w = csv.writer(sys.stdout, lineterminator="\n")
        w.writerow(["Date", "Time", "Item", "Quantity", "Unit",
                    "Value (%s)" % currency, "Reason", "Note", "Sent by",
                    "Picture", "Entry id"])
        for e in rows:
            w.writerow([e.get("d", ""), e.get("t", ""), e.get("item", ""),
                        e.get("qty", ""), e.get("unit", ""),
                        "" if e.get("price") in (None, "") else e.get("price"),
                        e.get("reason", ""), e.get("note", ""), e.get("by", ""),
                        "yes" if e.get("photo") else ("expired" if e.get("hadPhoto") else "no"),
                        e.get("id", "")])
        sys.stderr.write("%d row(s) for %s, generated %s Abu Dhabi time\n"
                         % (len(rows), "every day on file" if args.all else day, gen))
        return 0 if rows else 9

    wb = Workbook()

    # ------------------------------------------------------------ report --
    ws = wb.active
    ws.title = "Wastage"
    ws["A1"] = "SHAN VILLAGE - DAILY WASTAGE"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = ("Every day on file" if args.all else "%s" % day)
    ws["A2"].font = Font(size=12, color="756B58")
    ws["A3"] = "Generated %s Abu Dhabi time - values in %s" % (gen, currency)
    ws["A3"].font = Font(size=10, color="9C917A")

    header(ws, 5, ["Date", "Time", "Item", "Quantity", "Unit", "Reason",
                   "Value (%s)" % currency, "Sent by", "Note", "Picture"],
           [12, 8, 30, 11, 9, 20, 13, 18, 44, 10])
    r = 6
    total = 0.0
    priced = 0
    for e in rows:
        value = num(e.get("price"))
        if value is not None:
            total += value
            priced += 1
        vals = [e.get("d", ""), e.get("t", ""), e.get("item", ""),
                num(e.get("qty")), e.get("unit", ""), e.get("reason", ""),
                value, e.get("by", ""), e.get("note", ""),
                "yes" if e.get("photo") else ("expired" if e.get("hadPhoto") else "no")]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            if j in (1, 2, 4, 5, 7, 10):
                c.alignment = CENTER
            if j == 9:
                c.alignment = WRAP
            if j == 7 and value is not None:
                c.number_format = "#,##0.00"
                c.fill = MONEY_FILL
        r += 1

    if not rows:
        ws.cell(row=6, column=1, value="Nothing was recorded.").font = Font(color="756B58")
        r = 7

    r += 1
    ws.cell(row=r, column=3, value="Total").font = Font(bold=True)
    c = ws.cell(row=r, column=7, value=round(total, 2) if priced else None)
    c.font, c.fill, c.border, c.number_format = Font(bold=True), SUB_FILL, BORDER, "#,##0.00"
    ws.cell(row=r, column=8,
            value="%d of %d entries carried a value" % (priced, len(rows))).font = Font(
        size=10, color="756B58")
    ws.freeze_panes = "A6"
    if rows:
        ws.auto_filter.ref = "A5:J%d" % (5 + len(rows))

    # ------------------------------------------------------------ by item --
    def rollup(sheet_name, key, label):
        w = wb.create_sheet(sheet_name)
        agg = {}
        for e in rows:
            k = (e.get(key) or "(not given)").strip() or "(not given)"
            a = agg.setdefault(k, {"n": 0, "value": 0.0, "priced": 0, "qty": {}})
            a["n"] += 1
            v = num(e.get("price"))
            if v is not None:
                a["value"] += v
                a["priced"] += 1
            q, u = num(e.get("qty")), (e.get("unit") or "").strip()
            if q is not None and u:
                a["qty"][u] = a["qty"].get(u, 0) + q
        header(w, 1, [label, "Times", "Quantity", "Value (%s)" % currency, "Priced"],
               [30, 9, 22, 14, 9])
        i = 2
        for k in sorted(agg, key=lambda x: (-agg[x]["value"], -agg[x]["n"], x)):
            a = agg[k]
            qty = " + ".join("%g %s" % (round(v, 3), u) for u, v in sorted(a["qty"].items()))
            for j, v in enumerate([k, a["n"], qty, round(a["value"], 2) if a["priced"] else None,
                                   a["priced"]], start=1):
                c = w.cell(row=i, column=j, value=v)
                c.border = BORDER
                if j >= 2:
                    c.alignment = CENTER
                if j == 4 and v is not None:
                    c.number_format = "#,##0.00"
            i += 1
        if not agg:
            w.cell(row=2, column=1, value="Nothing was recorded.").font = Font(color="756B58")
        w.freeze_panes = "A2"

    rollup("By item", "item", "Item")
    rollup("By reason", "reason", "Reason")
    rollup("By person", "by", "Sent by")

    wb.save(out)
    print("wrote %s  (%d entries, %s %.2f across %d priced)"
          % (out, len(rows), currency, total, priced))


if __name__ == "__main__":
    sys.exit(main() or 0)
