#!/usr/bin/env python3
"""Build the stock-take report from the inventory page's state.

    python3 scripts/export-stocktake.py --state <artifact.html|state.json> \
        [--day 2026-08-25] [--all] [--csv] [--out "...xlsx"]

Stock is never stored on the page - it is the sum of the movements
ledger - so on-hand figures here are derived the same way the page
derives them, and cannot drift from what a counter saw.

--csv prints the rows on stdout instead of writing a workbook. Text that
size crosses into Google Drive intact; a binary workbook of any size does
not always. Exit 9 means there was nothing to report, so the caller can
skip an empty upload.
"""
import argparse, csv, datetime, json, re, sys

GULF = datetime.timezone(datetime.timedelta(hours=4))


def load_state(path):
    raw = open(path, encoding="utf8").read()
    m = re.search(r'<script id="state" type="application/json">(\{.*?)</script>', raw, re.S)
    state = json.loads(m.group(1) if m else raw)
    if "takes" not in state and "moves" not in state:
        sys.exit("no inventory state found in %s" % path)
    return state


def num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def n3(v):
    return round(num(v) + 0.0, 3)


def index(state):
    items = {i["id"]: i for i in state.get("items") or []}
    locs = {l["id"]: l for l in state.get("locations") or []}
    return items, locs


def on_hand(state):
    """item id -> location id -> quantity, from the movements ledger."""
    out = {}
    for m in state.get("moves") or []:
        out.setdefault(m.get("i"), {})
        loc = m.get("l")
        out[m["i"]][loc] = n3(out[m["i"]].get(loc, 0) + num(m.get("q")))
    return out


def take_rows(state, days):
    """One row per counted line, across the takes in range."""
    items, locs = index(state)
    rows = []
    for t in state.get("takes") or []:
        if days and t.get("date") not in days:
            continue
        counters = ", ".join(t.get("byNames") or []) or (t.get("by") or "")
        for item_id, L in (t.get("lines") or {}).items():
            if L.get("q") in (None, ""):
                continue
            it = items.get(item_id) or {}
            counted, book = n3(L.get("q")), n3(L.get("sys"))
            opening = bool(it.get("createdIn") and not num(L.get("sys")))
            var = 0.0 if opening else n3(counted - book)
            cost = num(L.get("cost"))
            rows.append({
                "ref": t.get("ref", ""),
                "date": t.get("date", ""),
                "location": (locs.get(t.get("loc")) or {}).get("name", t.get("loc", "")),
                "status": t.get("status", ""),
                "item": it.get("name", item_id),
                "sku": it.get("sku", ""),
                "counted": counted,
                "unit": L.get("unit") or it.get("unit") or "",
                "book": "" if opening else book,
                "variance": "opening" if opening else var,
                "value": "" if opening else round(var * cost, 2),
                "photo": "yes" if L.get("photo") else ("released" if L.get("hadPhoto") else "no"),
                "note": " ".join(x for x in [L.get("vcom"), L.get("note")] if x),
                "counted_by": L.get("byName") or counters,
            })
    rows.sort(key=lambda r: (r["date"], r["ref"], r["item"]))
    return rows


def stock_rows(state):
    items, locs = index(state)
    held = on_hand(state)
    rows = []
    for it in state.get("items") or []:
        if it.get("active") is False:
            continue
        by_loc = held.get(it["id"], {})
        total = n3(sum(by_loc.values()))
        rows.append({
            "item": it.get("name", ""),
            "sku": it.get("sku", ""),
            "category": it.get("cat", ""),
            "unit": it.get("unit", ""),
            "total": total,
            "by_loc": {(locs.get(k) or {}).get("name", k): v for k, v in by_loc.items()},
            "value": round(total * num(it.get("cost")), 2),
        })
    rows.sort(key=lambda r: r["item"].lower())
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--state", required=True)
    ap.add_argument("--day")
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--csv", action="store_true")
    ap.add_argument("--out")
    args = ap.parse_args()

    state = load_state(args.state)
    currency = state.get("cur") or "AED"
    day = args.day or datetime.datetime.now(GULF).strftime("%Y-%m-%d")
    days = None if args.all else {day}

    counted = take_rows(state, days)
    stock = stock_rows(state)

    if args.csv:
        w = csv.writer(sys.stdout, lineterminator="\n")
        w.writerow(["STOCK TAKE LINES"])
        w.writerow(["Reference", "Date", "Location", "Status", "Item", "SKU",
                    "Counted", "Unit", "Book", "Variance",
                    "Variance value (%s)" % currency, "Picture", "Note", "Counted by"])
        for r in counted:
            w.writerow([r["ref"], r["date"], r["location"], r["status"], r["item"],
                        r["sku"], r["counted"], r["unit"], r["book"], r["variance"],
                        r["value"], r["photo"], r["note"], r["counted_by"]])
        locs = sorted({k for r in stock for k in r["by_loc"]})
        w.writerow([])
        w.writerow(["STOCK ON HAND"])
        w.writerow(["Item", "SKU", "Category", "Unit"] + locs +
                   ["Total", "Value (%s)" % currency])
        for r in stock:
            w.writerow([r["item"], r["sku"], r["category"], r["unit"]] +
                       [r["by_loc"].get(l, 0) for l in locs] + [r["total"], r["value"]])
        sys.stderr.write("%d counted line(s) for %s, %d item(s) on hand\n"
                         % (len(counted), "every day on file" if args.all else day, len(stock)))
        return 0 if (counted or stock) else 9

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl is required for the workbook: pip install openpyxl "
                 "(or use --csv)")

    HEAD_FILL = PatternFill("solid", fgColor="171310")
    HEAD_FONT = Font(color="F7EFDD", bold=True, size=11)
    out = args.out or ("Shan Village Stock Take %s.xlsx"
                       % ("all days" if args.all else day))
    wb = Workbook()

    ws = wb.active
    ws.title = "Stock take"
    ws["A1"] = "SHAN VILLAGE - STOCK TAKE"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = "Every day on file" if args.all else day
    ws["A2"].font = Font(size=12, color="756B58")
    head = ["Reference", "Date", "Location", "Status", "Item", "SKU", "Counted",
            "Unit", "Book", "Variance", "Variance value (%s)" % currency,
            "Picture", "Note", "Counted by"]
    for i, v in enumerate(head, start=1):
        c = ws.cell(row=4, column=i, value=v)
        c.fill, c.font, c.alignment = HEAD_FILL, HEAD_FONT, Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(30, len(v) + 4))
    for r, row in enumerate(counted, start=5):
        for i, v in enumerate([row["ref"], row["date"], row["location"], row["status"],
                               row["item"], row["sku"], row["counted"], row["unit"],
                               row["book"], row["variance"], row["value"], row["photo"],
                               row["note"], row["counted_by"]], start=1):
            ws.cell(row=r, column=i, value=v)
    if not counted:
        ws.cell(row=5, column=1, value="No stock take was counted.").font = Font(color="756B58")

    ws2 = wb.create_sheet("Stock on hand")
    locs = sorted({k for row in stock for k in row["by_loc"]})
    head2 = ["Item", "SKU", "Category", "Unit"] + locs + ["Total", "Value (%s)" % currency]
    for i, v in enumerate(head2, start=1):
        c = ws2.cell(row=1, column=i, value=v)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        ws2.column_dimensions[get_column_letter(i)].width = max(10, min(30, len(str(v)) + 4))
    for r, row in enumerate(stock, start=2):
        vals = ([row["item"], row["sku"], row["category"], row["unit"]] +
                [row["by_loc"].get(l, 0) for l in locs] + [row["total"], row["value"]])
        for i, v in enumerate(vals, start=1):
            ws2.cell(row=r, column=i, value=v)

    wb.save(out)
    print("wrote %s - %d counted line(s), %d item(s) on hand"
          % (out, len(counted), len(stock)))
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
